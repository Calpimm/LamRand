import unittest
from lamrand.lamrand import LamRand
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference, ScatterChart, Series
import os
import statistics
import scipy.stats

class TestLamRandAdvanced(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.excel_file = 'test_results_advanced.xlsx'
        if not os.path.exists(cls.excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "TestResults"
            ws.append(["Test Name", "Result", "Details"])
            wb.save(cls.excel_file)

    def setUp(self):
        self.randomizer = LamRand()

    def save_to_excel(self, test_name, result, details=""):
        wb = load_workbook(self.excel_file)
        ws = wb.active
        ws.append([test_name, result, details])
        wb.save(self.excel_file)

    def test_next_float_distribution(self):
        values = [self.randomizer.next_float() for _ in range(100000)]
        mean = statistics.mean(values)
        variance = statistics.variance(values)
        ks_stat, p_value = scipy.stats.kstest(values, 'uniform', args=(0, 1))

        self.assertAlmostEqual(mean, 0.5, delta=0.01, msg="Mean is not approximately 0.5")
        self.assertAlmostEqual(variance, 1/12, delta=0.01, msg="Variance is not approximately 1/12")
        self.assertGreater(p_value, 0.05, "Kolmogorov-Smirnov test failed for uniform distribution")

        details = f"Mean: {mean}, Variance: {variance}, KS p-value: {p_value}"
        self.save_to_excel("test_next_float_distribution", "passed", details)
        print("✔ test_next_float_distribution passed")

    def test_next_int_distribution(self):
        values = [self.randomizer.next_int(1, 10) for _ in range(100000)]
        counts = [values.count(i) for i in range(1, 11)]
        chi2_stat, p_value = scipy.stats.chisquare(counts)

        self.assertTrue(all(9000 <= count <= 11000 for count in counts), "Values are not evenly distributed")
        self.assertGreater(p_value, 0.05, "Chi-square test failed for discrete uniform distribution")

        details = f"Counts: {counts}, Chi2 p-value: {p_value}"
        self.save_to_excel("test_next_int_distribution", "passed", details)
        print("✔ test_next_int_distribution passed")

    def test_next_gaussian_distribution(self):
        values = [self.randomizer.next_gaussian() for _ in range(100000)]
        mean = statistics.mean(values)
        stdev = statistics.stdev(values)
        ks_stat, p_value = scipy.stats.kstest(values, 'norm', args=(0, 1))

        self.assertAlmostEqual(mean, 0, delta=0.1, msg="Mean is not approximately 0")
        self.assertAlmostEqual(stdev, 1, delta=0.1, msg="Standard deviation is not approximately 1")
        self.assertGreater(p_value, 0.05, "Kolmogorov-Smirnov test failed for normal distribution")

        details = f"Mean: {mean}, Stdev: {stdev}, KS p-value: {p_value}"
        self.save_to_excel("test_next_gaussian_distribution", "passed", details)
        print("✔ test_next_gaussian_distribution passed")

    def test_next_bool_distribution(self):
        values = [self.randomizer.next_bool() for _ in range(100000)]
        true_count = values.count(True)
        false_count = values.count(False)
        chi2_stat, p_value = scipy.stats.chisquare([true_count, false_count])

        self.assertTrue(49000 <= true_count <= 51000, "True values are not evenly distributed")
        self.assertTrue(49000 <= false_count <= 51000, "False values are not evenly distributed")
        self.assertGreater(p_value, 0.05, "Chi-square test failed for boolean distribution")

        details = f"True: {true_count}, False: {false_count}, Chi2 p-value: {p_value}"
        self.save_to_excel("test_next_bool_distribution", "passed", details)
        print("✔ test_next_bool_distribution passed")

    def test_shuffle(self):
        data = [i for i in range(10)]
        shuffled_data = self.randomizer.shuffle(data[:])
        self.assertNotEqual(data, shuffled_data, "Data is not shuffled")
        self.assertCountEqual(data, shuffled_data, "Shuffled data does not have the same elements")

        self.save_to_excel("test_shuffle", "passed")
        print("✔ test_shuffle passed")

    def test_next_string(self):
        value = self.randomizer.next_string(10)
        self.assertIsInstance(value, str)
        self.assertEqual(len(value), 10)

        self.save_to_excel("test_next_string", "passed")
        print("✔ test_next_string passed")

    @classmethod
    def tearDownClass(cls):
        cls.create_charts()

    @classmethod
    def create_charts(cls):
        wb = load_workbook(cls.excel_file)
        ws = wb.active

        scatter_chart = ScatterChart()
        scatter_chart.title = "Test Results Over Time"
        scatter_chart.y_axis.title = "Result"
        scatter_chart.x_axis.title = "Test Case"

        data = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row, max_col=2)
        categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        scatter_chart.add_data(data, titles_from_data=True)
        scatter_chart.set_categories(categories)

        series = Series(data, categories, title="Results")
        scatter_chart.series.append(series)

        ws.add_chart(scatter_chart, "D1")
        wb.save(cls.excel_file)

if __name__ == '__main__':
    unittest.main()